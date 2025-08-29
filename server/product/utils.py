from datetime import datetime
from django.core.files.base import ContentFile
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins


def get_price_matrix(obj):
    from .models import Article, GenericProduct
    from django.db.models import Max
    from core.utils import int_to_code

    # Step 1: get latest created_at per (brand, product)
    latest_articles = (
        Article.objects.filter(generic_product__category=obj)
        .values("brand", "generic_product_id")
        .annotate(latest_created=Max("created_at"))
    )

    # Step 2: fetch those records
    articles = Article.objects.filter(
        generic_product__category=obj,
        created_at__in=[la["latest_created"] for la in latest_articles],
    ).select_related("generic_product")

    # Collect brands (strings)
    brands = sorted({a.brand for a in articles})

    # Header row
    rows = [["Product"] + brands]

    # collect all products in this category
    products = {
        gp.description: {b: "" for b in brands}
        for gp in GenericProduct.objects.filter(category=obj).order_by("description")
    }

    # fill with data from articles
    for art in articles:
        prod = art.generic_product.description
        products[prod][art.brand] = f"{art.selling_price}"
        # \n({int_to_code(art.purchase_price)})

    # Build rows
    for prod, brand_prices in products.items():
        rows.append([prod] + [brand_prices[b] for b in brands])

    return rows


def get_compatibility_matrix(obj):
    from .models import GenericProduct, Category, Motor

    # Only build matrix if this category has children
    child_cats = Category.objects.filter(parent_category=obj)
    if not child_cats.exists():
        return []

    # Collect all motors that appear under these child categories
    motors = Motor.objects.all()

    # Header row: "Motor" + each child category name
    rows = [["Motor"] + [cat.name for cat in child_cats]]

    # Build each row for a motor
    for motor in motors:
        row = [motor.model]
        for cat in child_cats:
            products = GenericProduct.objects.filter(
                category=cat, compatibility=motor
            ).values_list("description", flat=True)
            row.append(", ".join(products))
        rows.append(row)

    return rows


def generate_pricelist_assets(self):
    from core.utils import ReportBuilder

    timestamp = datetime.now().strftime("%y%m%d_%H%M")
    timestamp_head = (
        datetime.now().strftime("%B {day}, %Y").format(day=datetime.now().day)
    )

    rb = ReportBuilder(width=1200, height=800)
    rb.header(f"Pricelist for {self.name} ({timestamp_head})", size=25)
    rb.line(50)
    pricelist = get_price_matrix(self)
    cols = 1
    if len(pricelist) > 0:
        cols = len(pricelist[0])
    rb.table(
        pricelist,
        start=(50, 60),
        cell_size=(1100 / cols, 700 / len(pricelist)),
        font_size=0.95 * min(220 / cols, 700 / len(pricelist)),
    )

    buffer = BytesIO()
    rb.img.save(buffer, format="PNG")

    if self.pricelist_image:
        self.pricelist_image.delete(save=False)

    self.pricelist_image.save(
        f"pricelist_{self.pk}_{timestamp}.png",
        ContentFile(buffer.getvalue()),
        save=True,
    )

    # buffer = StringIO()
    # writer = csv.writer(buffer)
    # for row in pricelist:
    #     writer.writerow(row)

    if self.pricelist_file:
        self.pricelist_file.delete(save=False)

    # self.pricelist_file.save(
    #     f"pricelist_{self.pk}_{timestamp}.csv",
    #     ContentFile(buffer.getvalue().encode("utf-8")),
    #     save=True,
    # )
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricelist"

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

    total_width_units = 135  # ~A4 landscape width
    total_height_units = 520  # ~A4 landscape height

    cols = len(pricelist[0])
    rows = len(pricelist)

    col_width = total_width_units / cols
    row_height = total_height_units / rows
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # unlimited height (scale by width only)
    # Styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Apply equal col widths
    for c in range(1, cols + 1):
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = col_width

    # Apply equal row heights
    for r in range(1, rows + 1):
        ws.row_dimensions[r].height = row_height

    # Write data
    for r, row in enumerate(pricelist, start=1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(vertical="center", horizontal="center")
            cell.border = thin_border
            if r % 2 == 0:  # alternate row fill
                cell.fill = gray_fill

    # Save to memory
    buffer = BytesIO()
    wb.save(buffer)

    self.pricelist_file.save(
        f"pricelist_{self.pk}_{timestamp}.xlsx",
        ContentFile(buffer.getvalue()),
        save=True,
    )

    self.to_print_price = False
    self.save(update_fields=["to_print_price"])


def generate_compatibility_assets(self):
    from core.utils import ReportBuilder

    rb = ReportBuilder(width=600, height=400)
    rb.header(f"Compatibility for {self.name}")
    rb.line(70)
    rb.table([["Compatible With"], ["Motor A"], ["Motor B"]])

    buffer = BytesIO()
    rb.img.save(buffer, format="PNG")

    if self.compatibility_image:
        self.compatibility_image.delete(save=False)

    timestamp = datetime.now().strftime("%y%m%d_%H%M")

    self.compatibility_image.save(
        f"compatibility_{self.pk}_{timestamp}.png",
        ContentFile(buffer.getvalue()),
        save=True,
    )
    self.to_print_compatibility = False
    self.save(update_fields=["to_print_compatibility"])
