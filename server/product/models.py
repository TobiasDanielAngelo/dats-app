from my_django_app import fields
from django.core.exceptions import ValidationError
from io import BytesIO
from django.core.files.base import ContentFile
from .utils import get_compatibility_matrix, get_price_matrix
from datetime import datetime

import csv
from django.core.files.base import ContentFile
from io import StringIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins


class Unit(fields.CustomModel):
    name = fields.ShortCharField(display=True)


class Category(fields.CustomModel):
    parent_category = fields.SetNullOptionalForeignKey("self")
    name = fields.ShortCharField(display=True, unique=True)
    is_kit = fields.DefaultBooleanField(False)
    is_universal = fields.DefaultBooleanField(False)
    notes = fields.MediumCharField()
    to_print_price = fields.DefaultBooleanField(False)
    to_print_compatibility = fields.DefaultBooleanField(False)
    pricelist_image = fields.ImageField("prices")
    compatibility_image = fields.ImageField("compatibility")
    pricelist_file = fields.FileField("prices-csv")
    compatibility_file = fields.FileField("compatibility-csv")

    def generate_pricelist_image(self):
        print("Boom")
        from core.utils import ReportBuilder

        timestamp = datetime.now().strftime("%y%m%d_%H%M")
        timestamp_head = (
            datetime.now().strftime("%B {day}, %Y").format(day=datetime.now().day)
        )

        rb = ReportBuilder(width=1200, height=800)
        rb.header(f"Pricelist for {self.name} ({timestamp_head})", size=25)
        rb.line(50)
        pricelist = get_price_matrix(self)
        cols = 1
        if len(pricelist) > 0:
            cols = len(pricelist[0])
        rb.table(
            pricelist,
            start=(50, 60),
            cell_size=(1100 / cols, 700 / len(pricelist)),
            font_size=0.95 * min(220 / cols, 700 / len(pricelist)),
        )

        buffer = BytesIO()
        rb.img.save(buffer, format="PNG")

        if self.pricelist_image:
            self.pricelist_image.delete(save=False)

        self.pricelist_image.save(
            f"pricelist_{self.pk}_{timestamp}.png",
            ContentFile(buffer.getvalue()),
            save=True,
        )

        # buffer = StringIO()
        # writer = csv.writer(buffer)
        # for row in pricelist:
        #     writer.writerow(row)

        if self.pricelist_file:
            self.pricelist_file.delete(save=False)

        # self.pricelist_file.save(
        #     f"pricelist_{self.pk}_{timestamp}.csv",
        #     ContentFile(buffer.getvalue().encode("utf-8")),
        #     save=True,
        # )
        wb = Workbook()
        ws = wb.active
        ws.title = "Pricelist"

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

        total_width_units = 135  # ~A4 landscape width
        total_height_units = 520  # ~A4 landscape height

        cols = len(pricelist[0])
        rows = len(pricelist)

        col_width = total_width_units / cols
        row_height = total_height_units / rows
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # unlimited height (scale by width only)
        # Styles
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        gray_fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )

        # Apply equal col widths
        for c in range(1, cols + 1):
            col_letter = ws.cell(row=1, column=c).column_letter
            ws.column_dimensions[col_letter].width = col_width

        # Apply equal row heights
        for r in range(1, rows + 1):
            ws.row_dimensions[r].height = row_height

        # Write data
        for r, row in enumerate(pricelist, start=1):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.border = thin_border
                if r % 2 == 0:  # alternate row fill
                    cell.fill = gray_fill

        # Save to memory
        buffer = BytesIO()
        wb.save(buffer)

        self.pricelist_file.save(
            f"pricelist_{self.pk}_{timestamp}.xlsx",
            ContentFile(buffer.getvalue()),
            save=True,
        )

        self.to_print_price = False
        self.save(update_fields=["to_print_price"])

    def generate_compatibility_image(self):
        from core.utils import ReportBuilder

        rb = ReportBuilder(width=600, height=400)
        rb.header(f"Compatibility for {self.name}")
        rb.line(70)
        rb.table([["Compatible With"], ["Motor A"], ["Motor B"]])

        buffer = BytesIO()
        rb.img.save(buffer, format="PNG")

        if self.compatibility_image:
            self.compatibility_image.delete(save=False)

        timestamp = datetime.now().strftime("%y%m%d_%H%M")

        self.compatibility_image.save(
            f"compatibility_{self.pk}_{timestamp}.png",
            ContentFile(buffer.getvalue()),
            save=True,
        )
        self.to_print_compatibility = False
        self.save(update_fields=["to_print_compatibility"])


class Maker(fields.CustomModel):
    name = fields.ShortCharField(display=True, unique=True)


class Motor(fields.CustomModel):
    maker = fields.SetNullOptionalForeignKey(Maker)
    model = fields.ShortCharField(display=True, unique=True)


class GenericProduct(fields.CustomModel):
    category = fields.SetNullOptionalForeignKey(Category, display=True)
    compatibility = fields.OptionalManyToManyField(Motor, display=True)
    description = fields.MediumCharField(display=True)
    reorder_level = fields.LimitedIntegerField(1, 100000, 10)

    def clean(self):
        super().clean()
        motors = getattr(self, "_prefetched_compatibility", None)
        if motors is None:
            motors = self.compatibility.all() if self.pk else []

        # Normalize empties
        has_cat = bool(self.category)
        has_desc = bool(self.description)
        has_motor = bool(motors)

        if not (has_cat or has_desc or has_motor):
            return  # all empty, skip

        # Check for each motor if applicable, else handle no motor case
        check_motors = motors if has_motor else [None]

        for motor in check_motors:
            qs = GenericProduct.objects.exclude(pk=self.pk)

            # Build filter based on non-empty fields
            filters = {}
            if has_cat:
                filters["category"] = self.category
            if has_desc:
                filters["description"] = self.description
            if motor:
                filters["compatibility"] = motor

            if filters and qs.filter(**filters).exists():
                parts = []
                if has_cat:
                    parts.append(f"category '{self.category}'")
                if has_desc:
                    parts.append(f"description '{self.description}'")
                if motor:
                    parts.append(f"motor '{motor}'")

                msg = f"Combination of {', '.join(parts)} already exists."
                raise ValidationError(
                    {
                        "category": msg if has_cat else None,
                        "description": msg if has_desc else None,
                        "compatibility": msg if motor else None,
                    }
                )


class Article(fields.CustomModel):
    parent_article = fields.CascadeOptionalForeignKey("self")
    generic_product = fields.CascadeRequiredForeignKey(GenericProduct, display=True)
    brand = fields.ShortCharField(display=True)
    is_orig = fields.DefaultBooleanField(False, display=True)
    unit = fields.SetNullOptionalForeignKey(Unit)
    quantity_per_unit = fields.LimitedIntegerField(1, 100, 1)
    purchase_price = fields.LimitedDecimalField(0, 999999.99)
    selling_price = fields.LimitedDecimalField(0, 999999.99)

    def clean(self):
        super().clean()

        if self.purchase_price >= self.selling_price:
            raise ValidationError(
                {
                    "purchase_price": f"Must be less than selling price.",
                    "selling_price": f"Must be greater than purchase price.",
                }
            )


class Barcode(fields.CustomModel):
    product = fields.OneToOneField(Article, display=True)
    code = fields.ShortCharField(unique=True)


class PrintJob(fields.CustomModel):
    product = fields.CascadeRequiredForeignKey(Article, display=True)
    quantity = fields.LimitedIntegerField(1, 10000)


class ProductImage(fields.CustomModel):
    part = fields.CascadeRequiredForeignKey(Article, display=True)
    image = fields.ImageField(upload_to="part_images/")
    alt_text = fields.ShortCharField()


class Address(fields.CustomModel):
    name = fields.MediumCharField(display=True, unique=True)


class Location(fields.CustomModel):
    address = fields.CascadeRequiredForeignKey(Address, display=True)
    shelf = fields.ShortCharField(display=True)

    class Meta:
        unique_together = ("address", "shelf")
